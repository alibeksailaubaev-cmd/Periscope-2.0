#!/usr/bin/env python3
"""Генерация тестовых Excel-файлов для проверки дэшборда.

Создаёт в каталоге stand/ четыре файла в том же формате, что и рабочие выгрузки:
  act.xlsx         — акт приёма мойки (лист «ГРАФИК ДИНАМИКИ») для инкубации/ЗПП
  mb.xlsx          — свод рабочих журналов МБ (лист «Sheet1») по цеху инкубации
  act-broiler.xlsx — акт приёма мойки (лист «с марта 2025») для бройлерного цеха
  mb-broiler.xlsx  — свод МБ по бройлерному цеху

Требуется openpyxl:  pip install openpyxl
Запуск:  python3 dashboard/tests/make_fixtures.py
"""

import datetime
import pathlib
import random

from openpyxl import Workbook

OUT = pathlib.Path(__file__).resolve().parent / "stand"
BASE = datetime.datetime(2026, 7, 1)  # дэшборд отбрасывает всё, что раньше этой даты


def act_file(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ГРАФИК ДИНАМИКИ"
    ws.append(["Отчёт по приёмке", None, None])
    ws.append(["Дата", "Средний балл по мойке", "Комментарий"])
    rnd = random.Random(7)
    for i in range(14):
        ws.append([BASE + datetime.timedelta(days=i), round(rnd.uniform(7.2, 9.8), 1), "ok"])
    wb.save(path)


def mb_file(path, dept, zone, count=60, nok_every=7):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([
        "Дата отбора пробы", "Точка отбора пробы", "Результат",
        "Результат в допустимых пределах ОК или NOK", "Подразделение",
        "Корпус / Помещение", "Предприятие", "Группа материала", "Подгруппа материала",
    ])
    points = ["Стол сортировки", "Инкубатор №3", "Лоток", "Тележка", "Кран воды"]
    for i in range(count):
        nok = i % nok_every == 0
        ws.append([
            BASE + datetime.timedelta(days=i % 14),
            points[i % len(points)],
            "БГКП обнаружены" if nok else "КМАФАнМ 1x10^2",
            "NOK" if nok else "OK",
            dept, zone,
            "АО_Усть_Каменогорская_Птицефабрика", "Смывы", "Смыв_поверхности",
        ])
    wb.save(path)


def act_broiler_file(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "с марта 2025"
    ws.append([
        "Площадка/корпус", "Время суток", "Мастер мойки", "Дата", "Балл",
        "Контрольные точки (несоответствия)", "Несоответствие", "Приемка мойки", "Ответственный",
    ])
    rnd = random.Random(11)
    sites = ["Г5-1", "Г5-2", "Д3-1", "Е2-4", "БП12-3", "А1-2"]
    issues = ["Остатки пуха на поилках", "Загрязнение вентиляции", "Пыль на кормолиниях"]
    for i in range(40):
        score = rnd.choice([10, 10, 10, 9, 8, 7])
        ws.append([
            sites[i % len(sites)], rnd.choice(["День", "Ночь"]), f"Мастер {i % 3 + 1}",
            BASE + datetime.timedelta(days=i % 14), score,
            "" if score == 10 else "Поилки; Вентиляция",
            "" if score == 10 else rnd.choice(issues),
            "С первого раза" if score == 10 else "Перемывка", "Бригада мойки",
        ])
    wb.save(path)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    act_file(OUT / "act.xlsx")
    mb_file(OUT / "mb.xlsx", "Цех_инкубации", "Зона инкубации")
    act_broiler_file(OUT / "act-broiler.xlsx")
    mb_file(OUT / "mb-broiler.xlsx", "Цех_бройлерного_производства", "Г5-1", count=50, nok_every=8)
    print("готово:", ", ".join(p.name for p in sorted(OUT.glob("*.xlsx"))))


if __name__ == "__main__":
    main()
